import json
import logging
import os.path
from os import makedirs
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, fills
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.cell import get_column_letter

LOG_FILENAME = 'logs/cd-test-summary.log'
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

title_background = PatternFill(patternType=fills.FILL_SOLID,
                               start_color='00CAE3FF')


def generate_excel(source, output):
    logger.info(f"Selected source directory: {source}")
    logger.info(f"Output filename: {output}")

    seen_focus = set()
    test_cases = dict()

    # First pass to collect test cases and get the set of all seen focus areas
    p = Path(source)
    if not p.is_dir():
        raise Exception("The branch does not contain the requested folder")

    for f in list(p.glob('*.json')):
        logger.info(f"Starting to process {f} (type {type(f)}")

        key = f.parts[-1]
        test_cases[key] = list()

        with open(f, 'r', encoding='utf-8') as fx:
            contents = json.load(fx)

        for tc in contents:
            logger.debug(tc)
            test_cases[key].append(tc)
            if 'focus' in tc:
                seen_focus.update(tc['focus'])
            if 'dialog' in tc:
                for q in tc['dialog']:
                    if 'focus' in q:
                        seen_focus.update(q['focus'])
            if 'dataupload' in tc:
                for p in tc['dataupload']:
                    if 'focus' in p:
                        seen_focus.update(p['focus'])

    sorted_focus = sorted(list(seen_focus))
    logger.info(f"Seen focus areas: {sorted_focus}")
    logger.debug(test_cases)

    # Prepare workbook
    wb, ws, sorted_lookup = prepare_workbook(sorted_focus)

    # Second pass to write to the excel file
    row_index = 2
    for p in sorted(test_cases.keys()):
        logger.info(f"Starting detailed pass on {p}")
        tcs = test_cases[p]

        for tc in tcs:
            ws['A' + str(row_index)] = row_index
            ws['B' + str(row_index)] = p
            if 'appname' in tc: 
                ws['C' + str(row_index)] = tc['appname']
            else:
                ws['C' + str(row_index)] = "null"
            ws['D' + str(row_index)] = tc['language']

            # handle TTS cases where step 1 is TTS cmd and step 2 is ASR cmd
            if 'tts' in tc and 'query' in tc:
                logger.error(
                    f"Both tts and query are present in TC {tc} in {p}")
                exit(1)

            if 'tts' in tc:
                # Only show the voice for now
                q = f"TTS test using voice: {tc['tts'][0]['tts_voice']}"
                ws['E' + str(row_index)] = q

            if 'query' in tc:
                ws['E' + str(row_index)] = tc['query']

            audio = ''
            if 'audio' in tc:
                audio = 'True' if tc.get('audio') else ''

            disabled = ''
            if 'disabled' in tc:
                disabled = 'True' if tc.get('disabled') else ''

            # E and F are for NMAID and dictation type
            ws['F' + str(row_index)] = tc['nmaid']
            ws['G' + str(row_index)] = audio
            ws['H' + str(row_index)] = tc['dictationtype']
            ws['I' + str(row_index)] = disabled

            av = tc.get('verification', None)
            if av:
                for k, v in av.items():
                    sk = k.lower()
                    if "domain" in sk or "mbapps_slot_details.topic" in k or "ntg55_slot_details.searchType" in k:
                        ws['J' + str(row_index)] = v
                    elif "intention" in sk or "mbapps_slot_details.action" in k or "ntg55_slot_details.action" in k:
                        ws['K' + str(row_index)] = v
                    elif "mode" in k.lower():
                        ws['L' + str(row_index)] = v

            if 'focus' in tc:
                for fo in tc['focus']:
                    col_idx = sorted_lookup[fo]
                    c = ws.cell(row_index, col_idx, value="x")
                    c.alignment = Alignment(horizontal='center')

            if 'dialog' in tc:
                temp_row = row_index - 1
                for q in tc['dialog']:
                    temp_row += 1
                    ws['A' + str(temp_row)] = temp_row
                    if 'query' in q:
                        ws['E' + str(temp_row)] = q['query']
                    if 'focus' in q:
                        for fo in q['focus']:
                            col_idx = sorted_lookup[fo]
                            c = ws.cell(temp_row, col_idx, value="x")
                            c.alignment = Alignment(horizontal='center')
                    av = q.get('verification', None)
                    if av:
                        for k, v in av.items():
                            sk = k.lower()
                            if "domain" in sk or "mbapps_slot_details.topic" in k or "ntg55_slot_details.searchType" in k:
                                ws['J' + str(temp_row)] = v
                            elif "intention" in sk or "mbapps_slot_details.action" in k or "ntg55_slot_details.action" in k:
                                ws['K' + str(temp_row)] = v
                            elif "mode" in k.lower():
                                ws['L' + str(temp_row)] = v
                row_index = temp_row

            if 'dataupload' in tc:
                temp_row = row_index - 1
                for q in tc['dataupload']:
                    temp_row += 1
                    ws['A' + str(temp_row)] = temp_row
                    if 'cmdname' in q:
                        if 'query' in q:
                            ws['E' + str(temp_row)] = "Cmdname: " + q[
                                'cmdname'] + ", Query: " + q['query']
                        else:
                            ws['E' +
                               str(temp_row)] = "Cmdname: " + q['cmdname']
                    if 'focus' in q:
                        for fo in q['focus']:
                            col_idx = sorted_lookup[fo]
                            c = ws.cell(temp_row, col_idx, value="x")
                            c.alignment = Alignment(horizontal='center')
                    av = q.get('verification', None)
                    if av:
                        for k, v in av.items():
                            sk = k.lower()
                            if "domain" in sk or "mbapps_slot_details.topic" in k or "ntg55_slot_details.searchType" in k:
                                ws['J' + str(temp_row)] = v
                            elif "intention" in sk or "mbapps_slot_details.action" in k or "ntg55_slot_details.action" in k:
                                ws['K' + str(temp_row)] = v
                            elif "mode" in k.lower():
                                ws['L' + str(temp_row)] = v
                row_index = temp_row

            logger.info(f"Processed row {row_index}")
            row_index += 1

    # Freeze pane and save excel
    c = ws['F2']
    ws.freeze_panes = c
    wb.save(output)


def prepare_workbook(sorted_focus):
    logger.info("Preparing output excel")
    wb = Workbook()
    ws = wb.active
    ws.title = "CD Test"

    # Convert to tuples for column width; merge with sorted focus
    s_t = [(x, 10) for x in sorted_focus]
    column_headers = [('Line', 5), ('File', 15), ('Application', 15), ('Language', 15),
                      ('Query', 50), ('NMAID', 25), ('Audio', 10), ('Dictation', 20),
                      ('Disabled', 10), ('Domain', 15), ('Intention', 15),
                      ('Mode', 20)] + s_t
    idx = 1

    # Use index for variable length focus fields
    sorted_lookup = dict()
    for item in column_headers:
        logger.debug(f"Processing item: {item}")
        c = ws.cell(1, idx, value=item[0])
        c.border = thin_border
        c.font = Font(bold=True)
        c.fill = title_background
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(idx)].width = item[1]
        sorted_lookup[item[0]] = idx
        logger.debug(f"Creating column {item} at column index {idx}")
        idx += 1

    logger.debug(f"Column index: {sorted_lookup}")
    return wb, ws, sorted_lookup


def clean_up_log():
    if os.path.exists(LOG_FILENAME):
        with open(LOG_FILENAME, 'w') as f:
            f.close()
    else:
        makedirs('logs', exist_ok=True)


def set_up_logger():
    l = logging.getLogger(__name__)
    l.setLevel(logging.DEBUG)
    fh = logging.FileHandler(LOG_FILENAME, encoding='UTF-8')
    fh.setLevel(logging.DEBUG)
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    fh.setFormatter(formatter)
    ch.setFormatter(formatter)
    l.addHandler(fh)
    l.addHandler(ch)
    return l


clean_up_log()
logger = set_up_logger()
