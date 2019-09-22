from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, fills
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.cell import get_column_letter

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

title_background = PatternFill(patternType=fills.FILL_SOLID,
                               start_color='00CAE3FF')

def merge_excel(files_to_merge, dest_filename):

    onlyfiles = list(files_to_merge.glob('*.xlsx'))
    onlyfiles.sort()
    wb = Workbook()

    for f in onlyfiles:
        wbt = load_workbook(f)
        wst = wbt.worksheets[0]
        titles = Path(f).stem.split("+")
        if len(titles[0]) + len(titles[1]) > 29:
            ws = wb.create_sheet(F"{titles[0]} {titles[1][0]}")
        else:
            ws = wb.create_sheet(F"{titles[0]} {titles[1]}")
        for row in wst.iter_rows(min_row=1):
            for cell in row:
                ws[cell.coordinate].value = cell.value
                if cell.has_style:
                    ws[cell.coordinate].font = copy(cell.font)
                    ws[cell.coordinate].border = copy(cell.border)
                    ws[cell.coordinate].fill = copy(cell.fill)
                    ws[cell.coordinate].alignment = copy(cell.alignment)
        for idx, rd in wst.row_dimensions.items():
            ws.row_dimensions[idx] = copy(rd)
        for idx, cd in wst.column_dimensions.items():
            ws.column_dimensions[idx] = copy(cd)
        c = ws['E2']
        ws.freeze_panes = c

    wsf = wb.worksheets[0]
    wsf.title = "Index"
    column_headers = [('Branch', 20), ('Folder', 20), ('Link', 15)]
    idx = 1
    for item in column_headers:
            c = wsf.cell(1, idx, value=item[0])
            c.border = thin_border
            c.font = Font(bold=True)
            c.fill = title_background
            c.alignment = Alignment(horizontal='center')
            wsf.column_dimensions[get_column_letter(idx)].width = item[1]
            idx += 1
    idx = 2
    for w in wb.worksheets:
        if not w.title == "Index":
            titles = w.title.split(" ")
            wb.worksheets[0][f"A{idx}"].value = f"{titles[0]}"
            wb.worksheets[0][f"B{idx}"].value = f"{titles[1]}"
            wb.worksheets[0][f"C{idx}"].hyperlink = f"#'{w.title}'!A1"
            wb.worksheets[0][f"C{idx}"].value = "Click here"
            wb.worksheets[0][f"C{idx}"].style = "Hyperlink"
            idx += 1
    c = wsf['D2']
    wsf.freeze_panes = c

    wb.save(dest_filename)
