import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

import merge

DIRNAME = tempfile.TemporaryDirectory()


def test_merge_excel():

    temp_dir = Path(DIRNAME.name)

    name1 = temp_dir / "test1.xlsx"
    name2 = temp_dir / "test2.xlsx"
    name3 = temp_dir / "result.xlsx"

    wb1 = Workbook()
    wb1.save(name1)

    wb2 = Workbook()
    wb2.save(name2)

    merge.merge_excel(temp_dir, name3)
    wb3 = load_workbook(name3)

    assert wb3.worksheets[0].title == "test1"
    assert wb3.worksheets[0].max_row == wb1.worksheets[0].max_row
    assert wb3.worksheets[0].max_column == wb1.worksheets[0].max_column
    assert wb3.worksheets[1].title == "test2"
    assert wb3.worksheets[1].max_row == wb2.worksheets[0].max_row
    assert wb3.worksheets[1].max_column == wb2.worksheets[0].max_column
    assert len(wb3.sheetnames) == 2
