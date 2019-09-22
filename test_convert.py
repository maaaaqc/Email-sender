import json
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

import convert

DIRNAME = tempfile.TemporaryDirectory()


def test_generate_excel():
    temp_dir = Path(DIRNAME.name)
    out_name = temp_dir / "test.json"
    out_excel = temp_dir / "test.xlsx"
    sample = [{
        "domain": "a",
        "appkey": "b",
        "appname": "c",
        "nmaid": "d",
        "language": "e",
        "location": "f",
        "dictationtype": "g",
        "query": "h",
        "host": "i",
        "verification": {
            "$..actions[?(@.type == 'nlu_results')].Instances[0].nlu_classification.Domain":
            "j",
            "$..actions[?(@.type == 'nlu_results')].Instances[0].nlu_classification.Intention":
            "k"
        },
        "focus": ["l", "m"]
    }]
    with open(str(out_name), 'w', encoding='utf-8') as outfile:
        (json.dump(sample, outfile, ensure_ascii=False, indent=4))
    outfile.close()
    convert.generate_excel(temp_dir, out_excel)
    wb = load_workbook(out_excel)
    ws = wb.worksheets[0]

    assert ws["A2"].value == "test.json"
    assert ws["B2"].value == "c"
    assert ws["C2"].value == "e"
    assert ws["D2"].value == "h"
    assert ws["E2"].value == "d"
    assert ws["F2"].value == "g"
    assert ws["H2"].value == "j"
    assert ws["I2"].value == "k"


# def test_prepare_workbook():
